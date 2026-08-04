from datetime import datetime
from io import BytesIO

from services.negative_keyword_service import format_negative_keyword


def generate_negative_keyword_workbook(result: dict, context: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl. Install project dependencies.") from exc

    workbook = Workbook()
    workbook.creator = "Audilysis Negative Keyword Agent"

    search_sheet = workbook.active
    search_sheet.title = "Search Terms"
    negative_sheet = workbook.create_sheet("Negative Keywords")
    copy_sheet = workbook.create_sheet("Copy-Paste Ready")
    summary_sheet = workbook.create_sheet("Summary")

    rows = result["rows"]
    negatives = result["negative_keywords"]
    summary = result["summary"]
    report_date = context.get("report_date") or datetime.now().strftime("%Y-%m-%d")
    account_name = context.get("account_name") or "Uploaded Report"
    company_name = context.get("company_name") or context.get("brand_name") or "Audilysis"
    campaign_names = sorted({row["campaign"] for row in rows})

    build_search_terms_sheet(search_sheet, rows, report_date, account_name, campaign_names, Font, PatternFill, Alignment)
    build_negative_keywords_sheet(negative_sheet, negatives, report_date, Font, PatternFill, Alignment)
    build_copy_paste_sheet(copy_sheet, negatives, Font, PatternFill, Alignment)
    build_summary_sheet(summary_sheet, summary, report_date, account_name, company_name, campaign_names, Font, PatternFill)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_search_terms_sheet(sheet, rows, report_date, account_name, campaigns, Font, PatternFill, Alignment):
    blue = "1F4E79"
    sheet.merge_cells("A1:I1")
    sheet["A1"] = f"Google Ads Search Terms Report - {report_date}"
    sheet["A1"].font = Font(size=14, bold=True, color=blue)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:I2")
    sheet["A2"] = f"Account: {account_name} | Campaigns: {', '.join(campaigns)} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sheet["A2"].font = Font(size=10, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(horizontal="center")
    headers = ["#", "Search Term", "Campaign", "Clicks", "Impressions", "Cost ($)", "Conversions", "CTR (%)", "Avg CPC ($)"]
    write_header(sheet, 4, headers, blue, Font, PatternFill, Alignment)
    for index, row in enumerate(rows, start=1):
        avg_cpc = (row["cost"] or 0) / row["clicks"] if row["clicks"] else 0
        values = [
            index,
            row["search_term"],
            row["campaign"],
            row["clicks"] or 0,
            row["impressions"] or 0,
            row["cost"] or 0,
            row["conversions"] or 0,
            round((row["ctr"] or 0) * 100, 2),
            round(avg_cpc, 2),
        ]
        sheet.append(values)
        if index % 2 == 1:
            fill_row(sheet[index + 4], "EBF5FB", PatternFill)
        sheet[index + 4][5].number_format = "$#,##0.00"
        sheet[index + 4][8].number_format = "$#,##0.00"
    total_row = len(rows) + 5
    sheet[f"A{total_row}"] = "TOTALS"
    sheet[f"A{total_row}"].font = Font(bold=True)
    sheet[f"D{total_row}"] = sum(row["clicks"] or 0 for row in rows)
    sheet[f"E{total_row}"] = sum(row["impressions"] or 0 for row in rows)
    sheet[f"F{total_row}"] = sum(row["cost"] or 0 for row in rows)
    sheet[f"F{total_row}"].number_format = "$#,##0.00"
    sheet[f"G{total_row}"] = sum(row["conversions"] or 0 for row in rows)
    set_widths(sheet, [5, 50, 35, 10, 14, 12, 12, 10, 12])


def build_negative_keywords_sheet(sheet, negatives, report_date, Font, PatternFill, Alignment):
    red = "C0392B"
    wasted = sum(row["cost"] or 0 for row in negatives)
    sheet.merge_cells("A1:J1")
    sheet["A1"] = f"Audilysis Recommended Negative Keywords - {report_date}"
    sheet["A1"].font = Font(size=14, bold=True, color=red)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:J2")
    sheet["A2"] = f"{len(negatives)} suggestions | Estimated wasted spend: ${wasted:.2f} | Audilysis"
    sheet["A2"].font = Font(size=10, italic=True, color="666666")
    sheet["A2"].alignment = Alignment(horizontal="center")
    headers = ["#", "Search Term", "Campaign", "Negative Keyword", "Match Type", "Reason", "Confidence", "Risk", "Scope", "Cost ($)"]
    write_header(sheet, 4, headers, red, Font, PatternFill, Alignment)
    fills = {"High": "FADBD8", "Medium": "FEF9E7", "Low": "D5F5E3"}
    for index, row in enumerate(negatives, start=1):
        sheet.append([
            index,
            row["search_term"],
            row["campaign"],
            format_negative_keyword(row["negative_keyword"], row["match_type"]),
            row["match_type"],
            row["reason"],
            row["confidence"].upper(),
            row["risk"],
            row["scope"],
            row["cost"] or 0,
        ])
        fill_row(sheet[index + 4], fills.get(row["confidence"], "FEF9E7"), PatternFill)
        sheet[index + 4][9].number_format = "$#,##0.00"
    set_widths(sheet, [5, 45, 30, 35, 12, 50, 14, 12, 14, 12])


def build_copy_paste_sheet(sheet, negatives, Font, PatternFill, Alignment):
    blue = "1565C0"
    headers = ["Campaign", "Negative Keyword", "Match Type"]
    write_header(sheet, 1, headers, blue, Font, PatternFill, Alignment)
    seen = set()
    for row in negatives:
        keyword = format_negative_keyword(row["negative_keyword"], row["match_type"])
        key = (row["campaign"], keyword, row["match_type"])
        if key in seen:
            continue
        seen.add(key)
        sheet.append([row["campaign"], keyword, row["match_type"]])
    set_widths(sheet, [30, 40, 12])


def build_summary_sheet(sheet, summary, report_date, account_name, company_name, campaigns, Font, PatternFill):
    rows = [
        ("Report Date", report_date),
        ("Account", account_name),
        ("Company", company_name),
        ("Campaigns", ", ".join(campaigns)),
        ("Search Terms", summary["total_search_terms"]),
        ("Clicks", summary["total_clicks"]),
        ("Impressions", summary["total_impressions"]),
        ("Spend", f"${summary['total_cost']:.2f}"),
        ("Conversions", summary["total_conversions"]),
        ("Negatives", summary["negative_count"]),
        ("Review", summary["review_count"]),
        ("Keep", summary["keep_count"]),
        ("HIGH", summary["high_confidence"]),
        ("MEDIUM", summary["medium_confidence"]),
        ("LOW", summary["low_confidence"]),
        ("Estimated Wasted Spend", f"${summary['estimated_wasted_spend']:.2f}"),
    ]
    for index, (label, value) in enumerate(rows, start=1):
        sheet.append([label, value])
        sheet.cell(index, 1).font = Font(bold=True)
        if index % 2 == 1:
            fill_row(sheet[index], "D5F5E3", PatternFill)
    set_widths(sheet, [28, 70])


def write_header(sheet, row_number: int, headers: list[str], color: str, Font, PatternFill, Alignment):
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_number, index)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fill_row(row, color: str, PatternFill):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=color)


def set_widths(sheet, widths: list[int]):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
