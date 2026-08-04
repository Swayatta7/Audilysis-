import csv
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from agents.llm_client import openai_chat_completion
from agents.runtime_config import get_env_value
from db.storage import (
    create_negative_keyword_rule,
    delete_negative_keyword_rule,
    get_negative_keyword_rule_by_id,
    get_negative_keyword_instructions,
    get_negative_keyword_rules,
    get_scoped_negative_keyword_rule,
    reorder_negative_keyword_rules,
    set_negative_keyword_instructions,
    update_negative_keyword_rule,
)


MAX_UPLOAD_BYTES = 8 * 1024 * 1024
MAX_ROWS = 5000
MAX_LLM_ROWS_PER_BATCH = 40
ALLOWED_EXTENSIONS = {".csv", ".tsv", ".xlsx"}

SEARCH_TERM_COLUMNS = {"search term", "search terms", "query", "search query", "term", "search_term"}
CAMPAIGN_COLUMNS = {"campaign", "campaign name", "campaign_name"}
AD_GROUP_COLUMNS = {"ad group", "ad group name", "ad_group", "ad_group_name"}
CLICKS_COLUMNS = {"clicks", "click"}
IMPRESSIONS_COLUMNS = {"impressions", "impr.", "impr", "impression"}
COST_COLUMNS = {"cost", "cost ($)", "cost usd", "cost_usd", "cost_micros", "cost micros"}
MICROS_COST_COLUMNS = {"cost_micros", "cost micros"}
CONVERSIONS_COLUMNS = {"conversions", "conv.", "conv", "conversion"}
CTR_COLUMNS = {"ctr", "ctr (%)", "click through rate", "click-through rate"}

DEFAULT_RULES = [
    {
        "name": "Job seeker intent",
        "terms": ["job", "jobs", "career", "careers", "salary", "salaries", "internship", "intern", "resume", "vacancy", "hiring"],
        "classification": "NEGATIVE",
        "reason": "Job seeker intent.",
        "confidence": "HIGH",
        "risk": "LOW",
        "match_type": "PHRASE",
        "priority": 900,
        "enabled": True,
    },
    {
        "name": "Learning intent",
        "terms": ["course", "courses", "tutorial", "training", "pdf", "template", "meaning", "definition", "what is", "how to", "learn"],
        "classification": "NEGATIVE",
        "reason": "Learning or informational intent.",
        "confidence": "HIGH",
        "risk": "LOW",
        "match_type": "PHRASE",
        "priority": 800,
        "enabled": True,
    },
    {
        "name": "Price-sensitive intent",
        "terms": ["free", "cheap", "coupon", "discount", "promo code", "crack", "nulled"],
        "classification": "NEGATIVE",
        "reason": "Price-sensitive or free-seeking intent.",
        "confidence": "HIGH",
        "risk": "LOW",
        "match_type": "PHRASE",
        "priority": 700,
        "enabled": True,
    },
    {
        "name": "Low commercial intent",
        "terms": ["near me", "examples", "sample", "samples", "ideas", "book", "books"],
        "classification": "REVIEW",
        "reason": "Low commercial intent.",
        "confidence": "MEDIUM",
        "risk": "MEDIUM",
        "match_type": "EXACT",
        "priority": 600,
        "enabled": True,
    },
]

ALLOWED_CLASSIFICATIONS = {"KEEP", "REVIEW", "NEGATIVE"}
ALLOWED_CONFIDENCE = {"HIGH", "MEDIUM", "LOW"}
ALLOWED_RISK = {"HIGH", "MEDIUM", "LOW"}
ALLOWED_MATCH_TYPES = {"EXACT", "PHRASE", "BROAD"}


class NegativeKeywordError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.status_code = status_code


def ensure_rule_access(owner_key: str, user_id: int | None, rule_id: int):
    scoped_rule = get_scoped_negative_keyword_rule(rule_id, user_id, owner_key)
    if scoped_rule:
        return scoped_rule
    if get_negative_keyword_rule_by_id(rule_id):
        raise NegativeKeywordError("You do not have access to modify this rule.", status_code=403)
    raise NegativeKeywordError("Rule not found.", status_code=404)


@dataclass
class SearchTermRow:
    search_term: str
    campaign: str
    ad_group: str
    clicks: int | None
    impressions: int | None
    cost: float | None
    conversions: float | None
    ctr: float | None
    source_row: int
    raw: dict


def parse_search_terms_upload(file_storage) -> tuple[list[SearchTermRow], dict]:
    if not file_storage:
        raise NegativeKeywordError("Upload a Google Ads search-term report.")

    filename = Path(file_storage.filename or "").name
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise NegativeKeywordError("Upload a CSV, TSV, or XLSX search-term report.")

    content = file_storage.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise NegativeKeywordError("The uploaded report is too large. Keep it under 8 MB.")
    if not content:
        raise NegativeKeywordError("The uploaded report is empty.")

    if extension == ".xlsx":
        rows = _parse_xlsx(content)
    else:
        rows = _parse_delimited(content, delimiter="\t" if extension == ".tsv" else ",")

    normalized = normalize_search_term_rows(rows)
    metadata = {
        "filename": filename,
        "file_type": extension.lstrip("."),
        "source_rows": len(rows),
        "parsed_rows": len(normalized),
    }
    return normalized, metadata


def _parse_delimited(content: bytes, delimiter: str) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise NegativeKeywordError("The uploaded report has no header row.")
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise NegativeKeywordError("XLSX parsing requires openpyxl. Install project dependencies.") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    header_row = None
    headers = None
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [clean_header(value) for value in row]
        if any(value in SEARCH_TERM_COLUMNS for value in values):
            header_row = index
            headers = values
            break
    if not headers:
        raise NegativeKeywordError("Could not find a header row containing a Search Term column.")

    parsed = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value not in (None, "") for value in row):
            continue
        parsed.append({
            headers[column_index] or f"column_{column_index + 1}": value
            for column_index, value in enumerate(row)
            if column_index < len(headers)
        } | {"__source_row": row_index})
    return parsed


def normalize_search_term_rows(rows: Iterable[dict]) -> list[SearchTermRow]:
    normalized = []
    for fallback_index, row in enumerate(rows, start=2):
        lookup = {clean_header(key): value for key, value in row.items()}
        search_term = clean_text(first_value(lookup, SEARCH_TERM_COLUMNS))
        if not search_term:
            continue
        raw_cost = first_value(lookup, COST_COLUMNS)
        cost_is_micros = any(name in lookup for name in MICROS_COST_COLUMNS)
        normalized.append(SearchTermRow(
            search_term=search_term,
            campaign=clean_text(first_value(lookup, CAMPAIGN_COLUMNS)) or "Unassigned Campaign",
            ad_group=clean_text(first_value(lookup, AD_GROUP_COLUMNS)),
            clicks=parse_int(first_value(lookup, CLICKS_COLUMNS)),
            impressions=parse_int(first_value(lookup, IMPRESSIONS_COLUMNS)),
            cost=parse_money(raw_cost, micros=cost_is_micros),
            conversions=parse_float(first_value(lookup, CONVERSIONS_COLUMNS)),
            ctr=parse_ctr(first_value(lookup, CTR_COLUMNS)),
            source_row=int(row.get("__source_row") or fallback_index),
            raw={str(key): value for key, value in row.items() if not str(key).startswith("__")},
        ))
        if len(normalized) > MAX_ROWS:
            raise NegativeKeywordError("The uploaded report has too many rows to process safely.")

    if not normalized:
        raise NegativeKeywordError("No valid search terms were found in the uploaded report.")
    return normalized


def ensure_negative_keyword_defaults():
    existing = get_negative_keyword_rules()
    if existing:
        return existing
    for rule in DEFAULT_RULES:
        create_negative_keyword_rule(None, "__system__", rule)
    return get_negative_keyword_rules()


def get_negative_keyword_workspace_state(owner_key: str | None = None, user_id: int | None = None) -> dict:
    scoped_owner_key = owner_key or "__system__"
    if user_id is None and scoped_owner_key == "__system__":
        rules = ensure_negative_keyword_defaults()
    else:
        rules = get_negative_keyword_rules(user_id, scoped_owner_key)
        if not rules:
            for rule in ensure_negative_keyword_defaults():
                cloned = {key: value for key, value in rule.items() if key not in {"id", "owner_key", "created_at", "updated_at"}}
                create_negative_keyword_rule(user_id, scoped_owner_key, cloned)
            rules = get_negative_keyword_rules(user_id, scoped_owner_key)
    settings = get_negative_keyword_instructions(user_id, scoped_owner_key)
    return {
        "rules": rules,
        "custom_instructions": settings.get("custom_instructions") or "",
        "instructions_updated_at": settings.get("updated_at"),
    }


def create_rule(owner_key: str, user_id: int | None, payload: dict) -> dict:
    rule = normalize_rule_payload(payload)
    create_negative_keyword_rule(user_id, owner_key, rule)
    return get_negative_keyword_workspace_state(owner_key, user_id)


def update_rule(owner_key: str, user_id: int | None, rule_id: int, payload: dict) -> dict:
    ensure_rule_access(owner_key, user_id, rule_id)
    rule = normalize_rule_payload(payload, partial=True)
    update_negative_keyword_rule(user_id, owner_key, rule_id, rule)
    return get_negative_keyword_workspace_state(owner_key, user_id)


def delete_rule(owner_key: str, user_id: int | None, rule_id: int) -> dict:
    ensure_rule_access(owner_key, user_id, rule_id)
    delete_negative_keyword_rule(user_id, owner_key, rule_id)
    return get_negative_keyword_workspace_state(owner_key, user_id)


def reorder_rules(owner_key: str, user_id: int | None, rule_ids: list[int]) -> dict:
    reorder_negative_keyword_rules(user_id, owner_key, [int(rule_id) for rule_id in rule_ids])
    return get_negative_keyword_workspace_state(owner_key, user_id)


def save_custom_instructions(owner_key: str, user_id: int | None, custom_instructions: str) -> dict:
    set_negative_keyword_instructions(user_id, owner_key, clean_text(custom_instructions))
    return get_negative_keyword_workspace_state(owner_key, user_id)


def normalize_rule_payload(payload: dict, partial: bool = False) -> dict:
    normalized = {}
    fields = {
        "name": clean_text(payload.get("name")),
        "terms": parse_terms(payload.get("terms")),
        "classification": normalize_label(payload.get("classification"), ALLOWED_CLASSIFICATIONS),
        "reason": clean_text(payload.get("reason")),
        "confidence": normalize_label(payload.get("confidence"), ALLOWED_CONFIDENCE),
        "risk": normalize_label(payload.get("risk"), ALLOWED_RISK),
        "match_type": normalize_label(payload.get("match_type"), ALLOWED_MATCH_TYPES),
    }
    for key, value in fields.items():
        if partial and payload.get(key) is None:
            continue
        if key == "terms":
            if not partial or payload.get(key) is not None:
                if not value:
                    raise NegativeKeywordError("Rules must include at least one term.")
                normalized[key] = value
            continue
        if not value:
            raise NegativeKeywordError(f"Rule field '{key}' is required.")
        normalized[key] = value
    if "enabled" in payload:
        normalized["enabled"] = bool(payload.get("enabled"))
    if "priority" in payload and payload.get("priority") not in (None, ""):
        normalized["priority"] = int(payload.get("priority"))
    return normalized


def analyze_search_terms(rows: list[SearchTermRow], context: dict, workspace_state: dict | None = None) -> dict:
    workspace = workspace_state or get_negative_keyword_workspace_state(
        context.get("_owner_key"),
        context.get("_user_id"),
    )
    company = clean_text(context.get("company_name") or context.get("brand_name"))
    target_locations = parse_terms(context.get("target_locations"))
    excluded_locations = parse_terms(context.get("excluded_locations"))
    competitors = parse_terms(context.get("competitor_terms"))
    custom_negative_terms = parse_terms(context.get("custom_negative_terms"))
    conversion_threshold = safe_float(context.get("conversion_threshold"), 1.0)
    high_cost_threshold = safe_float(context.get("high_cost_threshold"), 100.0)
    min_clicks_for_spend_rule = max(1, int(safe_float(context.get("min_clicks_for_spend_rule"), 1)))

    active_rules = [rule for rule in workspace["rules"] if rule.get("enabled")]
    contextual_rules = build_contextual_rules(custom_negative_terms, competitors, excluded_locations)
    ordered_rules = sorted(contextual_rules + active_rules, key=lambda item: (int(item.get("priority", 0)), int(item.get("id", 0))), reverse=True)

    analyzed = []
    for row in rows:
        analyzed.append(classify_row(
            row,
            company=company,
            target_locations=target_locations,
            excluded_locations=excluded_locations,
            conversion_threshold=conversion_threshold,
            high_cost_threshold=high_cost_threshold,
            min_clicks_for_spend_rule=min_clicks_for_spend_rule,
            ordered_rules=ordered_rules,
        ))

    llm_status = "skipped"
    llm_prompt_preview = ""
    llm_error = ""
    llm_api_key = get_env_value("OPENAI_API_KEY")
    if llm_api_key:
        llm_rows, llm_status, llm_prompt_preview, llm_error = maybe_refine_with_llm(
            analyzed,
            context=context,
            ordered_rules=ordered_rules,
            custom_instructions=workspace.get("custom_instructions") or "",
            api_key=llm_api_key,
        )
        if llm_rows:
            analyzed = llm_rows

    negatives = [item for item in analyzed if item["classification"] == "NEGATIVE"]
    reviews = [item for item in analyzed if item["classification"] == "REVIEW"]
    keeps = [item for item in analyzed if item["classification"] == "KEEP"]
    total_cost = sum(item["cost"] or 0 for item in analyzed)
    wasted_cost = sum(item["cost"] or 0 for item in negatives)

    return {
        "summary": {
            "total_search_terms": len(analyzed),
            "negative_count": len(negatives),
            "review_count": len(reviews),
            "keep_count": len(keeps),
            "total_clicks": sum(item["clicks"] or 0 for item in analyzed),
            "total_impressions": sum(item["impressions"] or 0 for item in analyzed),
            "total_cost": round(total_cost, 2),
            "total_conversions": round(sum(item["conversions"] or 0 for item in analyzed), 2),
            "estimated_wasted_spend": round(wasted_cost, 2),
            "high_confidence": len([item for item in negatives if item["confidence"] == "HIGH"]),
            "medium_confidence": len([item for item in negatives if item["confidence"] == "MEDIUM"]),
            "low_confidence": len([item for item in negatives if item["confidence"] == "LOW"]),
        },
        "rows": analyzed,
        "negative_keywords": negatives,
        "review_terms": reviews,
        "keep_terms": keeps,
        "copy_paste_lines": build_copy_paste_lines(negatives),
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "active_rules": serialize_rules_for_response(ordered_rules),
        "custom_instructions": workspace.get("custom_instructions") or "",
        "llm_status": llm_status,
        "llm_error": llm_error,
        "llm_prompt_preview": llm_prompt_preview,
    }


def build_contextual_rules(custom_negative_terms: list[str], competitors: list[str], excluded_locations: list[str]) -> list[dict]:
    rules = []
    if custom_negative_terms:
        rules.append({
            "id": 0,
            "name": "Always negative terms",
            "terms": custom_negative_terms,
            "classification": "NEGATIVE",
            "reason": "Matched custom negative term rule.",
            "confidence": "HIGH",
            "risk": "LOW",
            "match_type": "PHRASE",
            "priority": 1200,
            "enabled": True,
        })
    if competitors:
        rules.append({
            "id": 0,
            "name": "Competitor terms",
            "terms": competitors,
            "classification": "REVIEW",
            "reason": "Competitor or irrelevant brand search; review before excluding.",
            "confidence": "MEDIUM",
            "risk": "MEDIUM",
            "match_type": "PHRASE",
            "priority": 1100,
            "enabled": True,
        })
    if excluded_locations:
        rules.append({
            "id": 0,
            "name": "Excluded locations",
            "terms": excluded_locations,
            "classification": "NEGATIVE",
            "reason": "Location mismatch.",
            "confidence": "HIGH",
            "risk": "LOW",
            "match_type": "PHRASE",
            "priority": 1000,
            "enabled": True,
        })
    return rules


def classify_row(row: SearchTermRow, **context) -> dict:
    term_lower = row.search_term.lower()
    reasons = []
    matched_rule = ""
    risk = "LOW"
    confidence = "LOW"
    classification = "KEEP"
    match_type = ""
    keyword = ""
    scope = "Campaign"

    conversions = row.conversions
    clicks = row.clicks or 0
    cost = row.cost or 0.0
    ctr = row.ctr

    if conversions is not None and conversions >= context["conversion_threshold"]:
        reasons.append("Term has recorded conversions; do not add as a negative without manual review.")
        classification = "KEEP"
        risk = "HIGH"
        confidence = "HIGH"
    else:
        matched = first_matching_rule(term_lower, context["ordered_rules"])
        if matched:
            classification = matched["classification"]
            matched_rule = matched["name"]
            reasons.append(matched["reason"])
            confidence = matched["confidence"]
            risk = matched["risk"]
            match_type = matched["match_type"] if classification == "NEGATIVE" else ""
            keyword = recommended_keyword(row.search_term, matched) if classification == "NEGATIVE" else ""
        elif context["target_locations"] and any(location in term_lower for location in {"india", "noida", "delhi", "mumbai", "kolkata", "pakistan", "bangladesh"}):
            if not contains_any(term_lower, context["target_locations"]):
                classification = "NEGATIVE"
                matched_rule = "Location mismatch"
                reasons.append("Location mismatch.")
                confidence = "HIGH"
                risk = "LOW"
                match_type = "PHRASE"
                keyword = row.search_term.lower()
        elif cost >= context["high_cost_threshold"] and conversions in (None, 0):
            classification = "NEGATIVE"
            matched_rule = "High spend zero conversions"
            reasons.append("High spend with zero recorded conversions.")
            confidence = "HIGH"
            risk = "LOW"
            keyword = row.search_term
            match_type = "EXACT"
        elif clicks >= context["min_clicks_for_spend_rule"] and cost > 0 and conversions in (None, 0):
            classification = "REVIEW"
            matched_rule = "Spend review"
            reasons.append("Spend and clicks with zero recorded conversions; review before excluding.")
            confidence = "MEDIUM"
            risk = "MEDIUM"
        elif ctr is not None and row.impressions and row.impressions >= 50 and ctr < 0.01:
            classification = "REVIEW"
            matched_rule = "Low CTR review"
            reasons.append("Low CTR on meaningful impression volume.")
            confidence = "MEDIUM"
            risk = "MEDIUM"
        else:
            reasons.append("No strong negative keyword signal found.")
            confidence = "MEDIUM"

    if classification == "NEGATIVE":
        risk = negative_risk(keyword or row.search_term, row.search_term, conversions, current_risk=risk)

    return {
        "search_term": row.search_term,
        "classification": classification,
        "campaign": row.campaign,
        "campaign_id": str(row.raw.get("campaign_id") or ""),
        "ad_group": row.ad_group,
        "negative_keyword": keyword if classification == "NEGATIVE" else "",
        "match_type": match_type if classification == "NEGATIVE" else "",
        "reason": " ".join(reasons),
        "matched_rule": matched_rule,
        "confidence": confidence,
        "risk": risk,
        "scope": scope,
        "clicks": row.clicks,
        "impressions": row.impressions,
        "cost": row.cost,
        "conversions": row.conversions,
        "ctr": row.ctr,
        "source_row": row.source_row,
    }


def first_matching_rule(term_lower: str, ordered_rules: list[dict]) -> dict | None:
    for rule in ordered_rules:
        if not rule.get("enabled", True):
            continue
        terms = rule.get("terms") or []
        if contains_any(term_lower, terms):
            return rule
    return None


def recommended_keyword(search_term: str, rule: dict) -> str:
    text = re.sub(r"\s+", " ", search_term.lower()).strip()
    terms = sorted(rule.get("terms") or [], key=len, reverse=True)
    for token in terms:
        if token and re.search(rf"\b{re.escape(token)}\b", text):
            return token
        if token and token in text:
            return token
    return text


def negative_risk(keyword: str, search_term: str, conversions, current_risk: str = "LOW") -> str:
    if conversions and conversions > 0:
        return "HIGH"
    if current_risk == "HIGH":
        return current_risk
    if len(keyword.split()) <= 1 and keyword.lower() not in {"job", "jobs", "free", "cheap"}:
        return "MEDIUM"
    if keyword.lower() != search_term.lower() and len(keyword.split()) <= 2:
        return "MEDIUM"
    return current_risk or "LOW"


def maybe_refine_with_llm(analyzed_rows: list[dict], context: dict, ordered_rules: list[dict], custom_instructions: str, api_key: str):
    prompt_preview = build_llm_prompt(context, ordered_rules, custom_instructions, analyzed_rows[:3])[:600]
    batches = [analyzed_rows[index:index + MAX_LLM_ROWS_PER_BATCH] for index in range(0, len(analyzed_rows), MAX_LLM_ROWS_PER_BATCH)]
    refined = []
    for batch in batches:
        _response_payload, content = openai_chat_completion(
            api_key=api_key,
            prompt=build_llm_prompt(context, ordered_rules, custom_instructions, batch),
        )
        if not content:
            return None, "failed", prompt_preview, "LLM did not return structured content."
        try:
            payload = json.loads(content)
        except json.JSONDecodeError:
            return None, "failed", prompt_preview, "LLM returned invalid JSON."
        validated = validate_llm_payload(payload, batch, ordered_rules)
        if validated is None:
            return None, "failed", prompt_preview, "LLM output failed validation."
        refined.extend(validated)
    merged = merge_llm_rows(analyzed_rows, refined)
    return merged, "validated", prompt_preview, ""


def build_llm_prompt(context: dict, ordered_rules: list[dict], custom_instructions: str, batch: list[dict]) -> str:
    prompt_payload = {
        "task": "Classify Google Ads search terms as KEEP, REVIEW, or NEGATIVE using the active rules and business context. Return only JSON with a top-level key named results.",
        "business_context": {
            "company_name": context.get("company_name") or context.get("brand_name") or "",
            "account_name": context.get("account_name") or "",
            "target_locations": parse_terms(context.get("target_locations")),
            "excluded_locations": parse_terms(context.get("excluded_locations")),
            "competitor_terms": parse_terms(context.get("competitor_terms")),
            "custom_negative_terms": parse_terms(context.get("custom_negative_terms")),
        },
        "custom_instructions": custom_instructions or "",
        "active_rules": serialize_rules_for_response(ordered_rules),
        "allowed_values": {
            "classification": sorted(ALLOWED_CLASSIFICATIONS),
            "confidence": sorted(ALLOWED_CONFIDENCE),
            "risk": sorted(ALLOWED_RISK),
            "match_type": sorted(ALLOWED_MATCH_TYPES),
        },
        "rows": [
            {
                "source_row": row["source_row"],
                "search_term": row["search_term"],
                "campaign": row["campaign"],
                "clicks": row["clicks"],
                "impressions": row["impressions"],
                "cost": row["cost"],
                "conversions": row["conversions"],
                "ctr": row["ctr"],
                "baseline_classification": row["classification"],
                "baseline_reason": row["reason"],
                "baseline_matched_rule": row["matched_rule"],
            }
            for row in batch
        ],
        "response_contract": {
            "results": [
                {
                    "source_row": "number",
                    "classification": "KEEP | REVIEW | NEGATIVE",
                    "reason": "string",
                    "matched_rule": "string",
                    "confidence": "HIGH | MEDIUM | LOW",
                    "risk": "HIGH | MEDIUM | LOW",
                    "negative_keyword": "string, required only when classification is NEGATIVE",
                    "match_type": "EXACT | PHRASE | BROAD, required only when classification is NEGATIVE",
                }
            ]
        },
    }
    return json.dumps(prompt_payload, indent=2)


def validate_llm_payload(payload: dict, batch: list[dict], ordered_rules: list[dict]) -> list[dict] | None:
    rows_by_source = {row["source_row"]: row for row in batch}
    known_rule_names = {rule["name"] for rule in ordered_rules}
    results = payload.get("results")
    if not isinstance(results, list) or len(results) != len(batch):
        return None
    validated = []
    for item in results:
        source_row = int(item.get("source_row", 0))
        original = rows_by_source.get(source_row)
        if not original:
            return None
        classification = normalize_label(item.get("classification"), ALLOWED_CLASSIFICATIONS)
        confidence = normalize_label(item.get("confidence"), ALLOWED_CONFIDENCE)
        risk = normalize_label(item.get("risk"), ALLOWED_RISK)
        matched_rule = clean_text(item.get("matched_rule"))
        if classification not in ALLOWED_CLASSIFICATIONS or confidence not in ALLOWED_CONFIDENCE or risk not in ALLOWED_RISK:
            return None
        if matched_rule and matched_rule not in known_rule_names and matched_rule not in {
            "High spend zero conversions",
            "Spend review",
            "Low CTR review",
            "Location mismatch",
        }:
            return None
        negative_keyword = clean_text(item.get("negative_keyword"))
        match_type = normalize_label(item.get("match_type"), ALLOWED_MATCH_TYPES) if item.get("match_type") else ""
        if classification == "NEGATIVE" and (not negative_keyword or match_type not in ALLOWED_MATCH_TYPES):
            return None
        if classification != "NEGATIVE":
            negative_keyword = ""
            match_type = ""
        validated.append({
            **original,
            "classification": classification,
            "reason": clean_text(item.get("reason")) or original["reason"],
            "matched_rule": matched_rule or original["matched_rule"],
            "confidence": confidence,
            "risk": risk,
            "negative_keyword": negative_keyword,
            "match_type": match_type,
        })
    return validated


def merge_llm_rows(original_rows: list[dict], refined_rows: list[dict]) -> list[dict]:
    refined_by_source = {row["source_row"]: row for row in refined_rows}
    return [refined_by_source.get(row["source_row"], row) for row in original_rows]


def build_copy_paste_lines(negative_rows: list[dict]) -> list[str]:
    lines = []
    seen = set()
    for row in negative_rows:
        keyword = format_negative_keyword(row["negative_keyword"], row["match_type"])
        key = (row["campaign"], keyword, row["match_type"])
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"{row['campaign']}\t{keyword}\t{row['match_type']}")
    return lines


def build_negative_keyword_csv(rows: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    columns = [
        "classification",
        "search_term",
        "campaign",
        "campaign_id",
        "ad_group",
        "negative_keyword",
        "match_type",
        "matched_rule",
        "reason",
        "confidence",
        "risk",
        "clicks",
        "impressions",
        "cost",
        "conversions",
        "ctr",
    ]
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(column, "") for column in columns])
    return output.getvalue()


def format_negative_keyword(keyword: str, match_type: str) -> str:
    clean = clean_text(keyword)
    if not clean:
        return ""
    if match_type == "PHRASE":
        return f'"{clean}"'
    if match_type == "BROAD":
        return clean
    return f"[{clean}]"


def serialize_rules_for_response(rules: list[dict]) -> list[dict]:
    return [
        {
            "id": rule.get("id"),
            "name": rule.get("name"),
            "terms": rule.get("terms") or [],
            "classification": rule.get("classification"),
            "reason": rule.get("reason"),
            "confidence": rule.get("confidence"),
            "risk": rule.get("risk"),
            "match_type": rule.get("match_type"),
            "enabled": bool(rule.get("enabled", True)),
            "priority": int(rule.get("priority", 0)),
        }
        for rule in rules
    ]


def first_value(row: dict, names: set[str]):
    for name in names:
        if name in row:
            return row[name]
    return None


def clean_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def parse_terms(value) -> list[str]:
    if isinstance(value, list):
        raw = value
    else:
        raw = re.split(r"[\n,]", str(value or ""))
    return [item.strip().lower() for item in raw if item and item.strip()]


def contains_any(text: str, terms) -> bool:
    return any(term and term in text for term in terms)


def parse_int(value):
    parsed = parse_float(value)
    return int(parsed) if parsed is not None else None


def parse_money(value, micros: bool = False):
    parsed = parse_float(value)
    if parsed is None:
        return None
    if micros:
        return round(parsed / 1000000, 2)
    return round(parsed, 2)


def parse_ctr(value):
    parsed = parse_float(value)
    if parsed is None:
        return None
    return parsed / 100 if parsed > 1 else parsed


def parse_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace(",", "").replace("%", "").strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def safe_float(value, default: float) -> float:
    parsed = parse_float(value)
    return default if parsed is None else parsed


def normalize_label(value, allowed: set[str]) -> str:
    label = clean_text(value).upper().replace(" ", "_")
    if not label:
        return ""
    if label in allowed:
        return label
    return ""
