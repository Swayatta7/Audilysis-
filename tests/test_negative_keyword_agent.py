import io
import os
import sqlite3
import unittest
from unittest.mock import patch

from werkzeug.datastructures import FileStorage
from werkzeug.security import generate_password_hash

from app import app
from agents.negative_keyword import REPORT_DIR
from db.storage import DB_PATH, create_user, get_negative_keyword_instructions, get_negative_keyword_rules, get_user_by_email
from services.negative_keyword_report import generate_negative_keyword_workbook
from services.negative_keyword_service import (
    NegativeKeywordError,
    analyze_search_terms,
    build_negative_keyword_csv,
    get_negative_keyword_workspace_state,
    parse_search_terms_upload,
)
from services.ownership import build_owner_context


SAMPLE_CSV = """Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions,CTR (%)
seo jobs,Brand Campaign,12,100,45.50,0,12
best seo agency,Brand Campaign,8,200,75,2,4
free seo audit template,Lead Campaign,5,90,18,0,5.5
seo agency in pakistan,Lead Campaign,3,80,9,0,3.75
"""

XLSX_SAMPLE_HEADERS = ["Search Term", "Campaign", "Clicks", "Impressions", "Cost ($)", "Conversions", "CTR (%)"]


def reset_negative_keyword_tables():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM negative_keyword_rules")
    cursor.execute("DELETE FROM negative_keyword_settings_v2")
    cursor.execute("DELETE FROM negative_keyword_audit")
    cursor.execute("DELETE FROM negative_keyword_reports")
    conn.commit()
    conn.close()


def owner_key_for_client(client):
    with client.session_transaction() as flask_session:
        browser_session_id = flask_session.get("browser_session_id")
        user_id = flask_session.get("auth_user_id")
    if not browser_session_id:
        client.get("/api/negative-keywords/google-ads/status")
        with client.session_transaction() as flask_session:
            browser_session_id = flask_session.get("browser_session_id")
            user_id = flask_session.get("auth_user_id")
    return build_owner_context(user_id, browser_session_id, True).owner_key


class NegativeKeywordAgentTestCase(unittest.TestCase):
    def setUp(self):
        self.client = app.test_client()
        app.config["TESTING"] = True
        reset_negative_keyword_tables()
        user = get_user_by_email("negative-tests@example.com")
        if not user:
            user_id = create_user("negative-tests@example.com", generate_password_hash("Password123"))
            user = {"id": user_id, "email": "negative-tests@example.com"}
        self.csrf_token = "negative-csrf"
        with self.client.session_transaction() as flask_session:
            flask_session["auth_user_id"] = user["id"]
            flask_session["csrf_token"] = self.csrf_token
            flask_session["browser_session_id"] = "negative-tests-session"
        original_open = self.client.open

        def open_with_auth(*args, **kwargs):
            method = str(kwargs.get("method", "GET")).upper()
            if method in {"POST", "PUT", "PATCH", "DELETE"}:
                headers = kwargs.setdefault("headers", {})
                headers.setdefault("X-CSRF-Token", self.csrf_token)
                headers.setdefault("X-Requested-With", "XMLHttpRequest")
            return original_open(*args, **kwargs)

        self.client.open = open_with_auth

    def test_csv_upload_parses_and_analyzes_negative_keywords(self):
        rows, metadata = parse_search_terms_upload(file_from_bytes(SAMPLE_CSV.encode(), "terms.csv"))

        result = analyze_search_terms(rows, {
            "company_name": "Example",
            "target_locations": ["united states"],
            "excluded_locations": ["pakistan"],
            "high_cost_threshold": 100,
        })

        self.assertEqual(metadata["parsed_rows"], 4)
        self.assertEqual(result["summary"]["total_search_terms"], 4)
        self.assertEqual(result["summary"]["negative_count"], 3)
        self.assertEqual(result["summary"]["keep_count"], 1)
        self.assertTrue(any(item["negative_keyword"] == "job" or item["negative_keyword"] == "jobs" for item in result["negative_keywords"]))
        self.assertTrue(all(item["classification"] in {"KEEP", "REVIEW", "NEGATIVE"} for item in result["rows"]))

    def test_cost_micros_column_is_converted_to_currency_units(self):
        csv_bytes = b"search_term,campaign,cost_micros,clicks,conversions\nbad traffic,Campaign,2500000,1,0\n"
        rows, _metadata = parse_search_terms_upload(file_from_bytes(csv_bytes, "terms.csv"))
        self.assertEqual(rows[0].cost, 2.5)

    def test_invalid_upload_extension_is_rejected(self):
        with self.assertRaises(NegativeKeywordError):
            parse_search_terms_upload(file_from_bytes(b"Search Term\nseo jobs\n", "terms.txt"))

    def test_empty_upload_is_rejected(self):
        with self.assertRaises(NegativeKeywordError):
            parse_search_terms_upload(file_from_bytes(b"", "terms.csv"))

    def test_malformed_csv_without_header_is_rejected(self):
        with self.assertRaises(NegativeKeywordError):
            parse_search_terms_upload(file_from_bytes(b"seo jobs\nseo agency\n", "terms.csv"))

    def test_xlsx_upload_parses_successfully(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(XLSX_SAMPLE_HEADERS)
        sheet.append(["seo jobs", "Brand Campaign", 12, 100, 45.5, 0, 12])
        buffer = io.BytesIO()
        workbook.save(buffer)
        rows, metadata = parse_search_terms_upload(file_from_bytes(buffer.getvalue(), "terms.xlsx"))

        self.assertEqual(metadata["parsed_rows"], 1)
        self.assertEqual(rows[0].search_term, "seo jobs")

    def test_workbook_contains_expected_sheets(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        rows, _metadata = parse_search_terms_upload(file_from_bytes(SAMPLE_CSV.encode(), "terms.csv"))
        result = analyze_search_terms(rows, {"company_name": "Example"})
        workbook_bytes = generate_negative_keyword_workbook(result, {"company_name": "Example"})
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)

        self.assertEqual(workbook.sheetnames, ["Search Terms", "Negative Keywords", "Copy-Paste Ready", "Summary"])

    def test_rule_priority_persistence_and_reorder(self):
        initial = self.client.get("/api/negative-keywords/rules").get_json()
        self.assertGreaterEqual(len(initial["rules"]), 1)

        create_response = self.client.post("/api/negative-keywords/rules", json={
            "name": "Very cautious free terms",
            "terms": "free, cheap",
            "classification": "REVIEW",
            "reason": "Review before excluding free terms.",
            "confidence": "HIGH",
            "risk": "HIGH",
            "match_type": "PHRASE",
            "enabled": True,
            "priority": 5000,
        })
        self.assertEqual(create_response.status_code, 200)
        rules = create_response.get_json()["rules"]
        self.assertEqual(rules[0]["name"], "Very cautious free terms")

        rows, _metadata = parse_search_terms_upload(file_from_bytes(SAMPLE_CSV.encode(), "terms.csv"))
        result = analyze_search_terms(
            rows,
            {"company_name": "Example"},
            workspace_state=get_negative_keyword_workspace_state(owner_key=owner_key_for_client(self.client)),
        )
        free_row = next(item for item in result["rows"] if item["search_term"] == "free seo audit template")
        self.assertEqual(free_row["classification"], "REVIEW")
        self.assertEqual(free_row["matched_rule"], "Very cautious free terms")

        reorder = self.client.post("/api/negative-keywords/rules/reorder", json={
            "rule_ids": [rule["id"] for rule in reversed(rules)]
        })
        self.assertEqual(reorder.status_code, 200)
        reloaded_rules = get_negative_keyword_rules(owner_key=owner_key_for_client(self.client))
        self.assertEqual(reloaded_rules[0]["id"], reversed(rules).__iter__().__next__()["id"])

    def test_disable_and_enable_rule_persist_and_affect_analysis(self):
        create_response = self.client.post("/api/negative-keywords/rules", json={
            "name": "Pricing review",
            "terms": "pricing",
            "classification": "REVIEW",
            "reason": "Review pricing queries.",
            "confidence": "HIGH",
            "risk": "MEDIUM",
            "match_type": "PHRASE",
            "enabled": True,
            "priority": 9000,
        })
        self.assertEqual(create_response.status_code, 200)
        created_rule = next(rule for rule in create_response.get_json()["rules"] if rule["name"] == "Pricing review")

        rows, _metadata = parse_search_terms_upload(file_from_bytes(
            b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions,CTR (%)\npricing overview,Brand Campaign,0,10,0,0,0.5\n",
            "terms.csv",
        ))
        enabled_result = analyze_search_terms(
            rows,
            {"company_name": "Example"},
            workspace_state=get_negative_keyword_workspace_state(owner_key=owner_key_for_client(self.client)),
        )
        enabled_row = next(item for item in enabled_result["rows"] if item["search_term"] == "pricing overview")
        self.assertEqual(enabled_row["classification"], "REVIEW")

        disable_response = self.client.put(f"/api/negative-keywords/rules/{created_rule['id']}", json={"enabled": False})
        self.assertEqual(disable_response.status_code, 200)
        disabled_rule = next(rule for rule in disable_response.get_json()["rules"] if rule["id"] == created_rule["id"])
        self.assertFalse(disabled_rule["enabled"])

        persisted_rules = get_negative_keyword_rules(owner_key=owner_key_for_client(self.client))
        self.assertFalse(next(rule for rule in persisted_rules if rule["id"] == created_rule["id"])["enabled"])

        disabled_result = analyze_search_terms(
            rows,
            {"company_name": "Example"},
            workspace_state=get_negative_keyword_workspace_state(owner_key=owner_key_for_client(self.client)),
        )
        disabled_row = next(item for item in disabled_result["rows"] if item["search_term"] == "pricing overview")
        self.assertEqual(disabled_row["classification"], "KEEP")
        self.assertNotEqual(disabled_row["matched_rule"], "Pricing review")

        refresh_rules = self.client.get("/api/negative-keywords/rules").get_json()["rules"]
        self.assertFalse(next(rule for rule in refresh_rules if rule["id"] == created_rule["id"])["enabled"])

        enable_response = self.client.put(f"/api/negative-keywords/rules/{created_rule['id']}", json={"enabled": True})
        self.assertEqual(enable_response.status_code, 200)
        enabled_rule = next(rule for rule in enable_response.get_json()["rules"] if rule["id"] == created_rule["id"])
        self.assertTrue(enabled_rule["enabled"])

    def test_cross_user_rule_update_is_rejected(self):
        create_response = self.client.post("/api/negative-keywords/rules", json={
            "name": "Private rule",
            "terms": "jobs",
            "classification": "NEGATIVE",
            "reason": "Private rule.",
            "confidence": "HIGH",
            "risk": "LOW",
            "match_type": "PHRASE",
            "enabled": True,
            "priority": 5000,
        })
        rule_id = next(rule for rule in create_response.get_json()["rules"] if rule["name"] == "Private rule")["id"]

        second_client = app.test_client()
        second_csrf_token = "negative-csrf-3"
        second_user = get_user_by_email("negative-tests-3@example.com")
        if not second_user:
            user_id = create_user("negative-tests-3@example.com", generate_password_hash("Password123"))
            second_user = {"id": user_id, "email": "negative-tests-3@example.com"}
        with second_client.session_transaction() as flask_session:
            flask_session["auth_user_id"] = second_user["id"]
            flask_session["csrf_token"] = second_csrf_token
            flask_session["browser_session_id"] = "negative-tests-session-3"
        second_original_open = second_client.open

        def second_open_with_auth(*args, **kwargs):
            method = str(kwargs.get("method", "GET")).upper()
            if method in {"POST", "PUT", "PATCH", "DELETE"}:
                headers = kwargs.setdefault("headers", {})
                headers.setdefault("X-CSRF-Token", second_csrf_token)
                headers.setdefault("X-Requested-With", "XMLHttpRequest")
            return second_original_open(*args, **kwargs)

        second_client.open = second_open_with_auth
        response = second_client.put(f"/api/negative-keywords/rules/{rule_id}", json={"enabled": False})
        self.assertEqual(response.status_code, 403)
        self.assertIn("do not have access", response.get_json()["message"])

        persisted_rules = get_negative_keyword_rules(owner_key=owner_key_for_client(self.client))
        self.assertTrue(next(rule for rule in persisted_rules if rule["id"] == rule_id)["enabled"])

    def test_custom_instructions_persist_after_restart(self):
        response = self.client.post("/api/negative-keywords/instructions", json={
            "custom_instructions": "Prefer REVIEW for broad informational terms unless spend is very high."
        })
        self.assertEqual(response.status_code, 200)
        settings = get_negative_keyword_instructions(owner_key=owner_key_for_client(self.client))
        self.assertIn("Prefer REVIEW", settings["custom_instructions"])

        follow_up = self.client.get("/api/negative-keywords/instructions")
        self.assertEqual(follow_up.status_code, 200)
        self.assertIn("Prefer REVIEW", follow_up.get_json()["custom_instructions"])

    def test_saved_rules_and_instructions_are_included_in_llm_prompt(self):
        self.client.post("/api/negative-keywords/rules", json={
            "name": "Brand safety",
            "terms": "example brand",
            "classification": "KEEP",
            "reason": "Protect branded terms.",
            "confidence": "HIGH",
            "risk": "HIGH",
            "match_type": "EXACT",
            "enabled": True,
            "priority": 6000,
        })
        self.client.post("/api/negative-keywords/instructions", json={
            "custom_instructions": "Never over-block branded search terms."
        })
        rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nexample brand pricing,Brand,1,10,3,0\n", "terms.csv"))[0][0]
        ]

        captured = {}

        def fake_completion(api_key, prompt, model="gpt-4o-mini"):
            captured["prompt"] = prompt
            return {}, '{"results":[{"source_row":2,"classification":"KEEP","reason":"Brand safety check.","matched_rule":"Brand safety","confidence":"HIGH","risk":"HIGH","negative_keyword":"","match_type":""}]}'

        with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=False):
            with patch("services.negative_keyword_service.openai_chat_completion", side_effect=fake_completion):
                result = analyze_search_terms(
                    rows,
                    {"company_name": "Example"},
                    workspace_state=get_negative_keyword_workspace_state(owner_key=owner_key_for_client(self.client)),
                )

        self.assertEqual(result["llm_status"], "validated")
        self.assertIn("Never over-block branded search terms.", captured["prompt"])
        self.assertIn("Brand safety", captured["prompt"])
        self.assertEqual(result["rows"][0]["classification"], "KEEP")

    def test_upload_context_fields_flow_into_prompt_and_report_metadata(self):
        rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nexample competitor,Brand,1,10,3,0\n", "terms.csv"))[0][0]
        ]
        captured = {}

        def fake_completion(api_key, prompt, model="gpt-4o-mini"):
            captured["prompt"] = prompt
            return {}, '{"results":[{"source_row":2,"classification":"REVIEW","reason":"Competitor review.","matched_rule":"Competitor terms","confidence":"MEDIUM","risk":"MEDIUM","negative_keyword":"","match_type":""}]}'

        context = {
            "company_name": "Example Co",
            "account_name": "North America Account",
            "target_locations": ["united states", "canada"],
            "excluded_locations": ["pakistan"],
            "competitor_terms": ["competitor"],
            "custom_negative_terms": ["forbidden"],
        }

        with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}, clear=False):
            with patch("services.negative_keyword_service.openai_chat_completion", side_effect=fake_completion):
                result = analyze_search_terms(rows, context)

        workbook_bytes = generate_negative_keyword_workbook(result, context)
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
        summary_sheet = workbook["Summary"]

        self.assertIn('"company_name": "Example Co"', captured["prompt"])
        self.assertIn('"account_name": "North America Account"', captured["prompt"])
        self.assertIn('"target_locations": [', captured["prompt"])
        self.assertIn('"excluded_locations": [', captured["prompt"])
        self.assertIn('"competitor_terms": [', captured["prompt"])
        self.assertIn('"custom_negative_terms": [', captured["prompt"])
        self.assertEqual(summary_sheet["B2"].value, "North America Account")
        self.assertEqual(summary_sheet["B3"].value, "Example Co")

    def test_competitor_terms_default_to_review_not_negative(self):
        rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nacme competitor pricing,Brand,1,10,3,0\n", "terms.csv"))[0][0]
        ]
        result = analyze_search_terms(rows, {
            "company_name": "Example",
            "competitor_terms": ["competitor"],
        })
        self.assertEqual(result["rows"][0]["classification"], "REVIEW")
        self.assertEqual(result["rows"][0]["matched_rule"], "Competitor terms")

    def test_custom_negative_terms_force_negative(self):
        rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nforbidden phrase,Brand,1,10,3,0\n", "terms.csv"))[0][0]
        ]
        result = analyze_search_terms(rows, {
            "company_name": "Example",
            "custom_negative_terms": ["forbidden"],
        })
        self.assertEqual(result["rows"][0]["classification"], "NEGATIVE")
        self.assertEqual(result["rows"][0]["matched_rule"], "Always negative terms")

    def test_threshold_fields_change_analysis_result(self):
        high_spend_rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nexpensive query,Brand,10,100,150,0\n", "terms.csv"))[0][0]
        ]
        low_threshold = analyze_search_terms(high_spend_rows, {"company_name": "Example", "high_cost_threshold": 100})
        high_threshold = analyze_search_terms(high_spend_rows, {"company_name": "Example", "high_cost_threshold": 200})
        self.assertEqual(low_threshold["rows"][0]["classification"], "NEGATIVE")
        self.assertEqual(high_threshold["rows"][0]["classification"], "REVIEW")

        spend_review_rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nbroad query,Brand,2,100,10,0\n", "terms.csv"))[0][0]
        ]
        clicks_two = analyze_search_terms(spend_review_rows, {"company_name": "Example", "min_clicks_for_spend_rule": 2})
        clicks_three = analyze_search_terms(spend_review_rows, {"company_name": "Example", "min_clicks_for_spend_rule": 3})
        self.assertEqual(clicks_two["rows"][0]["classification"], "REVIEW")
        self.assertEqual(clicks_three["rows"][0]["classification"], "KEEP")

        conversion_rows = [
            parse_search_terms_upload(file_from_bytes(b"Search Term,Campaign,Clicks,Impressions,Cost ($),Conversions\nseo jobs,Brand,12,100,45.50,1\n", "terms.csv"))[0][0]
        ]
        threshold_one = analyze_search_terms(conversion_rows, {"company_name": "Example", "conversion_threshold": 1})
        threshold_two = analyze_search_terms(conversion_rows, {"company_name": "Example", "conversion_threshold": 2})
        self.assertEqual(threshold_one["rows"][0]["classification"], "KEEP")
        self.assertEqual(threshold_two["rows"][0]["classification"], "NEGATIVE")

    def test_csv_export_works(self):
        rows, _metadata = parse_search_terms_upload(file_from_bytes(SAMPLE_CSV.encode(), "terms.csv"))
        result = analyze_search_terms(rows, {"company_name": "Example"})
        csv_text = build_negative_keyword_csv(result["negative_keywords"])
        self.assertIn("classification,search_term,campaign", csv_text)
        self.assertIn("seo jobs", csv_text)

        response = self.client.post("/api/negative-keywords/export/csv", json={"rows": result["negative_keywords"]})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "text/csv")

    def test_run_agent_accepts_multipart_upload_and_report_download_works(self):
        response = self.client.post("/run-agent", data={
            "agent": "negative_keyword",
            "company_name": "Example",
            "target_locations": "United States",
            "excluded_locations": "Pakistan",
            "search_terms_file": (io.BytesIO(SAMPLE_CSV.encode()), "terms.csv"),
        }, content_type="multipart/form-data")

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertTrue(data["success"])
        self.assertEqual(data["agent"], "Negative Keyword Agent")
        self.assertIn("workspace_state", data["data"])
        report_filename = data["data"]["report_filename"]

        try:
            download = self.client.get(data["data"]["report_url"])
            self.assertEqual(download.status_code, 200)
            self.assertEqual(
                download.mimetype,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.assertTrue(download.data.startswith(b"PK"))
        finally:
            report_path = REPORT_DIR / report_filename
            if report_path.exists():
                os.remove(report_path)

    @patch("app.fetch_google_ads_search_terms")
    def test_manual_upload_and_google_ads_use_same_analysis_pipeline(self, search_terms_mock):
        upload_rows, _metadata = parse_search_terms_upload(file_from_bytes(SAMPLE_CSV.encode(), "terms.csv"))
        search_terms_mock.return_value = (
            upload_rows,
            {
                "customer_id": "1234567890",
                "campaign_ids": ["1111111111"],
                "date_start": "2026-07-01",
                "date_end": "2026-07-07",
                "parsed_rows": len(upload_rows),
                "source_rows": len(upload_rows),
                "file_type": "google_ads_api",
                "filename": "live_google_ads",
            },
        )
        live_response = self.client.post("/api/negative-keywords/analyse", json={
            "company_name": "Example",
            "customer_id": "1234567890",
            "campaign_ids": ["1111111111"],
            "start_date": "2026-07-01",
            "end_date": "2026-07-07",
        })
        upload_response = self.client.post("/run-agent", data={
            "agent": "negative_keyword",
            "company_name": "Example",
            "search_terms_file": (io.BytesIO(SAMPLE_CSV.encode()), "terms.csv"),
        }, content_type="multipart/form-data")

        self.assertEqual(live_response.status_code, 200)
        self.assertEqual(upload_response.status_code, 200)
        self.assertEqual(
            live_response.get_json()["data"]["summary"]["negative_count"],
            upload_response.get_json()["data"]["summary"]["negative_count"],
        )

    def test_run_agent_json_behavior_remains_unchanged_for_existing_agent(self):
        response = self.client.post("/run-agent", json={"agent": "keyword_clustering", "keyword_list": "seo audit\nseo checklist"})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["success"])

    def test_owner_scoped_rules_instructions_and_reports_are_isolated_between_clients(self):
        second_client = app.test_client()
        second_csrf_token = "negative-csrf-2"
        second_user = get_user_by_email("negative-tests-2@example.com")
        if not second_user:
            user_id = create_user("negative-tests-2@example.com", generate_password_hash("Password123"))
            second_user = {"id": user_id, "email": "negative-tests-2@example.com"}
        with second_client.session_transaction() as flask_session:
            flask_session["auth_user_id"] = second_user["id"]
            flask_session["csrf_token"] = second_csrf_token
            flask_session["browser_session_id"] = "negative-tests-session-2"
        second_original_open = second_client.open

        def second_open_with_auth(*args, **kwargs):
            method = str(kwargs.get("method", "GET")).upper()
            if method in {"POST", "PUT", "PATCH", "DELETE"}:
                headers = kwargs.setdefault("headers", {})
                headers.setdefault("X-CSRF-Token", second_csrf_token)
                headers.setdefault("X-Requested-With", "XMLHttpRequest")
            return second_original_open(*args, **kwargs)

        second_client.open = second_open_with_auth

        self.client.post("/api/negative-keywords/rules", json={
            "name": "Client One Rule",
            "terms": "jobs",
            "classification": "NEGATIVE",
            "reason": "Client one rule.",
            "confidence": "HIGH",
            "risk": "LOW",
            "match_type": "PHRASE",
            "enabled": True,
            "priority": 7000,
        })
        self.client.post("/api/negative-keywords/instructions", json={
            "custom_instructions": "Client one instructions."
        })

        second_client.post("/api/negative-keywords/rules", json={
            "name": "Client Two Rule",
            "terms": "cheap",
            "classification": "REVIEW",
            "reason": "Client two rule.",
            "confidence": "MEDIUM",
            "risk": "MEDIUM",
            "match_type": "PHRASE",
            "enabled": True,
            "priority": 7100,
        })
        second_client.post("/api/negative-keywords/instructions", json={
            "custom_instructions": "Client two instructions."
        })

        first_rules = self.client.get("/api/negative-keywords/rules").get_json()
        second_rules = second_client.get("/api/negative-keywords/rules").get_json()
        self.assertTrue(any(rule["name"] == "Client One Rule" for rule in first_rules["rules"]))
        self.assertFalse(any(rule["name"] == "Client Two Rule" for rule in first_rules["rules"]))
        self.assertTrue(any(rule["name"] == "Client Two Rule" for rule in second_rules["rules"]))
        self.assertFalse(any(rule["name"] == "Client One Rule" for rule in second_rules["rules"]))
        self.assertIn("Client one instructions.", first_rules["custom_instructions"])
        self.assertIn("Client two instructions.", second_rules["custom_instructions"])

        upload_response = self.client.post("/run-agent", data={
            "agent": "negative_keyword",
            "company_name": "Example",
            "search_terms_file": (io.BytesIO(SAMPLE_CSV.encode()), "terms.csv"),
        }, content_type="multipart/form-data")
        report_url = upload_response.get_json()["data"]["report_url"]
        self.assertEqual(self.client.get(report_url).status_code, 200)
        self.assertEqual(second_client.get(report_url).status_code, 404)


def file_from_bytes(content: bytes, filename: str) -> FileStorage:
    return FileStorage(stream=io.BytesIO(content), filename=filename)


if __name__ == "__main__":
    unittest.main()
